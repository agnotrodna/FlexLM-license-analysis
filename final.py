import matplotlib.pyplot as plt
import pandas as pd
from io import StringIO
import matplotlib.dates as mdates
import os
#######################CLEAN APPDATA#########################
appdatafile = '%s\\FlexLM License Analyst\\' %  os.environ['APPDATA']
if not os.path.exists(appdatafile):
    os.makedirs(appdatafile)

file_path = '%slicense_report.txt' % appdatafile
tabledatapath = '%stable_DATA.txt' % appdatafile
vendor_in_file = '(NANOSOFT)'
exceltable = 'output.xlsx'

def cleaner(path_to_file):
    if os.path.exists(path_to_file):
        os.remove(path_to_file)
    else:
        print("good")

cleaner(file_path)
cleaner(tabledatapath)
cleaner(exceltable)

path_to_log = "C:\\Program Files (x86)\\Nanosoft\\Nanosoft License Server\\flex.log"
########################MAIN#################################
def samoe_glavnoe():
    def parse_license_report(file_content):
        date_entries = defaultdict(lambda: defaultdict(set))
        current_date = None

        for line in file_content.split('\n'):
            line = line.strip()
            if line.startswith('----') and line.endswith('----'):
                # Это строка с датой
                date_part = line.strip('-')
                if len(date_part) == 10:  # Проверяем, что это дата (формат YYYY-MM-DD)
                    current_date = date_part
            elif current_date and vendor_in_file in line:
                # Это строка с записью о лицензии
                parts = line.split('"')
                if len(parts) >= 3:
                    product = parts[1]
                    user_part = parts[2].split('@')[0].strip()
                    if 'OUT:' in line:
                        # Добавляем пользователя для продукта на текущую дату
                        date_entries[current_date][product].add(user_part)

        return date_entries


    def format_output(date_entries):
        output = []
        for date in sorted(date_entries.keys()):
            products = date_entries[date]
            output.append(f"Дата: {date}")
            output.append("------------------------------")
            for product in sorted(products.keys()):
                users = products[product]
                user_list = ", ".join(sorted(users))
                output.append(f"{product}: {len(users)} , {user_list}")
            output.append("")  # Пустая строка между датами
        return "\n".join(output)


    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()



    parsed_data = parse_license_report(content)
    result = format_output(parsed_data)
    print(result)


    data = result
    file_for_table = open(tabledatapath, "w+", encoding='utf-8')
    file_for_table.write(data)
    file_for_table.close()


    # Парсинг данных
    def parse_license_data(data):
        license_data = []
        current_date = None

        for line in StringIO(data):
            line = line.strip()
            if not line:
                continue

            if line.startswith("Дата:"):
                date_str = line.split(': ')[1].strip()
                current_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            elif line.startswith("------------------------------"):
                continue
            else:
                if current_date is None:
                    continue

                # Разделяем строку на название лицензии и остальную часть
                parts = line.split(':', 1)
                if len(parts) < 2:
                    continue

                license_name = parts[0].strip()
                rest = parts[1].split(',')

                try:
                    count = int(rest[0].strip())
                except (ValueError, IndexError):
                    continue

                users = [u.strip() for u in rest[1:] if u.strip()]

                license_data.append({
                    'Date': current_date,
                    'License': license_name,
                    'Count': count,
                    'Users': ', '.join(users) if users else 'None'
                })

        return license_data


    # Преобразуем данные в DataFrame
    license_data = parse_license_data(data)
    df = pd.DataFrame(license_data)

    if df.empty:
        print("Нет данных для построения графика")
    else:
        # Создаем сводную таблицу для графика
        pivot_df = df.pivot_table(index='Date', columns='License', values='Count', aggfunc='sum').fillna(0)

        # Создаем график
        plt.figure(figsize=(12, 6))

        # Рисуем линии для каждого продукта
        for product in pivot_df.columns:
            plt.plot(pivot_df.index, pivot_df[product], marker='o', label=product, linewidth=1)

        # Настраиваем отображение дат
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gca().xaxis.set_major_locator(mdates.DayLocator())
        plt.gcf().autofmt_xdate()  # Наклон подписей дат

        # Добавляем подписи
        plt.title('Использование лицензий по дням')
        plt.xlabel('Дата')
        plt.ylabel('Количество использованных лицензий')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend(title='Продукты', bbox_to_anchor=(1.05, 1), loc='upper left')

        plt.tight_layout()
        plt.show()

######################INTERFACE############################
from main_window import *
from subproc import *
import tkinter as tk
import sys
from PyQt5.QtWidgets import QGraphicsBlurEffect

tk_root = tk.Tk()
tk_root.withdraw()

blur_effect = QGraphicsBlurEffect()
blur_effect.setBlurRadius(8)



def calendar_dataset(self):
    selected_date = self.calendarWidget.selectedDate()
    formatted_date = selected_date.toString("yyyy-MM-dd")
        # Читаем данные из файла
    try:
        with open(tabledatapath, 'r', encoding='utf-8') as file:
            data = file.read()
    except FileNotFoundError:
        QtWidgets.QMessageBox.critical(self, "Ошибка", "Файл данных не найден!")


        # Разбиваем данные по датам
    date_blocks = data.split('Дата: ')[1:]

        # Ищем блок данных для выбранной даты
    target_data = []
    for block in date_blocks:
        if block.startswith(formatted_date):
                # Нашли нужную дату, извлекаем данные
            lines = block.split('\n')[2:]  # Пропускаем строку с датой и разделитель
            for line in lines:
                if line.strip():  # Пропускаем пустые строки
                    parts = line.split(' , ')
                    if len(parts) == 2:
                        product_info = parts[0].split(': ')
                        product_name = product_info[0]
                        licenses_count = product_info[1]
                        users = parts[1]
                        target_data.append([product_name, licenses_count, users])
            break

        # Создаем модель для таблицы
    model = QtGui.QStandardItemModel()
    model.setHorizontalHeaderLabels(['Продукт', 'Кол-во лицензий', 'Пользователи'])

        # Заполняем модель данными
    for row in target_data:
        items = [QtGui.QStandardItem(field) for field in row]
        model.appendRow(items)

        # Устанавливаем модель в таблицу
    self.tableView.setModel(model)

        # Настраиваем отображение таблицы
    self.tableView.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
    self.tableView.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
    self.tableView.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)


def _clear_table(self):
    """Очищает таблицу."""
    model = QtGui.QStandardItemModel()
    model.setHorizontalHeaderLabels(['Продукт', 'Кол-во лицензий', 'Пользователи'])
    self.tableView.setModel(model)


def _fill_table(self, data):
    """Заполняет таблицу данными."""
    model = QtGui.QStandardItemModel()
    model.setHorizontalHeaderLabels(['Продукт', 'Кол-во лицензий', 'Пользователи'])

    for row in data:
        items = [QtGui.QStandardItem(field) for field in row]
        model.appendRow(items)

    self.tableView.setModel(model)
    self.tableView.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
    self.tableView.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
    self.tableView.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)


app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()
ui.tableView.setStyleSheet("""
    /* Основной стиль QTableView */
    QTableView {
        font-family: "Segoe UI", sans-serif;
        font-size: 14px;
        background-color: #F9F9F9;
        color: #333;
        gridline-color: #D0D0D0;
        border: 1px solid #C0C0C0;
        border-radius: 8px;
        outline: 0px; /* Убирает пунктир при фокусе */
    }

    /* Альтернативный цвет для строк */
    QTableView::item {
        padding: 6px;
        border-bottom: 1px solid #E0E0E0;
    }

    QTableView::item:selected {
        background-color: #4A90E2;
        color: black;
    }

    QTableView::item:hover {
        background-color: #F0F8FF;
    }

    /* Заголовок таблицы */
    QHeaderView::section {
        background-color: #E0E0E0;
        color: black;
        font-weight: bold;
        padding: 6px;
        border: none;
        border-right: 1px solid #C0C0C0;
    }

    QHeaderView::section:checked {
        background-color: #4A90E2;
        color: white;
    }

    /* Горизонтальный скроллбар */
    QScrollBar:horizontal {
        background: #F0F0F0;
        height: 12px;
        margin: 0px;
    }

    QScrollBar::handle:horizontal {
        background: #C0C0C0;
        min-width: 20px;
        border-radius: 6px;
    }

    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
        background: none;
    }

    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
        background: none;
    }

    /* Вертикальный скроллбар */
    QScrollBar:vertical {
        background: #F0F0F0;
        width: 12px;
        margin: 0px;
    }

    QScrollBar::handle:vertical {
        background: #C0C0C0;
        min-height: 20px;
        border-radius: 6px;
    }

    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
        background: none;
    }

    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
        background: none;
    }
""")

ui.graphicsView.setGraphicsEffect(blur_effect)

# Динамически добавляем метод в ui
ui.calendarWidget.selectionChanged.connect(lambda: calendar_dataset(ui))
ui.calendarWidget.hide()


ui.label.setText("VER. 1.4")


def open_click():
    if vendor_in_file == '(NANOSOFT)':
        ui.calendarWidget.setStyleSheet("""
    /* Основной стиль календаря */
    QCalendarWidget {
        font-family: "Segoe UI", sans-serif;
        font-size: 14px;
        border: 2px solid #A0A0A0;
        border-radius: 10px;
        background-color: #F9F9F9;
    }

    /* Стиль заголовка */
    QCalendarWidget QWidget#qt_calendar_navigationbar {
        background-color: #4A90E2;
        border-radius: 8px;
        color: white;
        font-weight: bold;
    }

    /* Кнопки переключения месяца */
    QCalendarWidget QToolButton {
        color: white;
        font-size: 16px;
        padding: 4px 8px;
        margin: 2px;
        border-radius: 6px;
        background-color: transparent;
    }

    QCalendarWidget QToolButton:hover {
        background-color: rgba(255, 255, 255, 30);
    }

    /* Выпадающий список месяцев и годов */
    QCalendarWidget QMenu {
        background-color: #FFFFFF;
        border: 1px solid #C0C0C0;
        border-radius: 6px;
    }

    QCalendarWidget QMenu::item {
        padding: 4px 16px;
    }

    QCalendarWidget QMenu::item:selected {
        background-color: #4A90E2;
        color: white;
    }

    /* Таблица дней */
    QCalendarWidget QTableView {
        border: 4px solid #D0D0D0;
        border-radius: 6px;
        selection-background-color: #4A90E2;
        selection-color: white;
        font-size: 14px;
    }

    /* Дни недели */
    QCalendarWidget QHeaderView::section {
        background-color: #E0E0E0;
        color: black;
        font-weight: bold;
        padding: 4px;
        border: 4px solid #D0D0D0;
        border-radius: 4px;
    }

    /* Сегодняшняя дата */
    QCalendarWidget QDateEdit {
        background-color: #FFF3CD;
        color: #333;
        border: 3px solid #CCC;
        border-radius: 6px;
    }

    /* Выбранный день */
    QCalendarWidget QAbstractItemView:enabled {
        selection-background-color: #4A90E2;
        selection-color: white;
    }
""")

    if vendor_in_file == '(CSOFT)':
     ui.calendarWidget.setStyleSheet("""
    /* Основной стиль календаря */
    QCalendarWidget {
        font-family: "Segoe UI", sans-serif;
        font-size: 14px;
        border: 2px solid #A0A0A0;
        border-radius: 10px;
        background-color: #F9F9F9;
    }

    /* Стиль заголовка */
    QCalendarWidget QWidget#qt_calendar_navigationbar {
        background-color: #cd5c5c;
        border-radius: 8px;
        color: white;
        font-weight: bold;
    }

    /* Кнопки переключения месяца */
    QCalendarWidget QToolButton {
        color: white;
        font-size: 16px;
        padding: 4px 8px;
        margin: 2px;
        border-radius: 6px;
        background-color: transparent;
    }

    QCalendarWidget QToolButton:hover {
        background-color: rgba(255, 255, 255, 30);
    }

    /* Выпадающий список месяцев и годов */
    QCalendarWidget QMenu {
        background-color: #FFFFFF;
        border: 1px solid #C0C0C0;
        border-radius: 6px;
    }

    QCalendarWidget QMenu::item {
        padding: 4px 16px;
    }

    QCalendarWidget QMenu::item:selected {
        background-color: #cd5c5c;
        color: white;
    }

    /* Таблица дней */
    QCalendarWidget QTableView {
        border: 4px solid #D0D0D0;
        border-radius: 6px;
        selection-background-color: #cd5c5c;
        selection-color: white;
        font-size: 14px;
    }

    /* Дни недели */
    QCalendarWidget QHeaderView::section {
        background-color: #E0E0E0;
        color: black;
        font-weight: bold;
        padding: 4px;
        border: 4px solid #D0D0D0;
        border-radius: 4px;
    }

    /* Сегодняшняя дата */
    QCalendarWidget QDateEdit {
        background-color: #cd5c5c;
        color: #333;
        border: 3px solid #CCC;
        border-radius: 6px;
    }

    /* Выбранный день */
    QCalendarWidget QAbstractItemView:enabled {
        selection-background-color: #cd5c5c;
        selection-color: white;
    }
""")
     
    open_file = path_to_log
    if open_file:
        main(open_file)
        with open(file_path, 'r', encoding='utf-8') as file:
            samoe_glavnoe()


ui.actionOpen.triggered.connect(open_click)
ui.actionOpen.triggered.connect(ui.calendarWidget.show)

def close_click():
    sys.exit(1)

ui.actionClose.triggered.connect(close_click)

def vendor_change_csoft():
    global path_to_log, vendor_in_file
    path_to_log = "C:\\Program Files (x86)\\CSoft\\CS License Server\\flex.log"
    ui.graphicsView.setStyleSheet("background-color: rgb(237, 28, 36);")
    vendor_in_file = '(CSOFT)'
   




def vendor_change_nano():
    global path_to_log, vendor_in_file
    path_to_log = "C:\\Program Files (x86)\\Nanosoft\\Nanosoft License Server\\flex.log"
    ui.graphicsView.setStyleSheet("background-color: rgb(0, 172, 226);")
    vendor_in_file = '(NANOSOFT)'



########################EXPORT2EXCEL######################################################
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tkinter import filedialog


def txt_to_excel():
    # Создаем новую книгу Excel
    print("НАЧАЛИ")
    wb = openpyxl.Workbook()

    # Читаем исходный файл
    with open(tabledatapath, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    current_date = None
    current_sheet = None

    for line in lines:
        line = line.strip()

        # Если строка начинается с "Дата:", создаем новый лист
        if line.startswith('Дата:'):
            current_date = line.split(': ')[1]
            current_sheet = wb.create_sheet(title=current_date)

            # Добавляем заголовки столбцов
            current_sheet['A1'] = 'Продукт'
            current_sheet['B1'] = 'Количество'
            current_sheet['C1'] = 'Пользователи'

            # Форматируем заголовки
            for col in range(1, 4):
                current_sheet[get_column_letter(col) + '1'].font = Font(bold=True)

            row_num = 2

        # Если строка содержит данные о программе
        elif line and '------------------------------' not in line and current_sheet:
            # Разделяем строку на части
            parts = line.split(' , ')
            program_info = parts[0].split(': ')
            program_name = program_info[0]
            program_count = program_info[1]
            users = parts[1] if len(parts) > 1 else ''

            # Записываем данные в текущий лист
            current_sheet[f'A{row_num}'] = program_name
            current_sheet[f'B{row_num}'] = program_count
            current_sheet[f'C{row_num}'] = users

            row_num += 1

    # Удаляем пустой лист, созданный по умолчанию
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Создаем окно tkinter для выбора места сохранения
    root = tk.Tk()
    root.withdraw()  # Скрываем основное окно

    # Запрашиваем у пользователя место сохранения и имя файла
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Сохранить файл как"
    )

    # Если пользователь не отменил диалог
    if file_path:
        # Сохраняем книгу
        wb.save(file_path)
        print(f'Файл {file_path} успешно создан.')
    else:
        print('Сохранение отменено пользователем.')



ui.action.triggered.connect(txt_to_excel)

ui.graph.triggered.connect(samoe_glavnoe)

ui.pushButton.clicked.connect(vendor_change_csoft)

ui.pushButton_2.clicked.connect(vendor_change_nano)

sys.exit(app.exec_())